import json
import re
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill


def now_iso():
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def text(value):
    return "" if value is None else str(value).strip()


def normalize(value):
    return "".join(text(value).upper().split())


def to_int(value, default=0):
    try:
        if value in (None, ""):
            return default
        return int(float(value))
    except Exception:
        return default


def to_float(value, default=0.0):
    try:
        if value in (None, ""):
            return default
        return float(value)
    except Exception:
        return default


def parse_bool(value):
    return text(value).lower() in {"1", "true", "yes", "y", "是"}


def bool_label(value):
    return "是" if value else "否"


def make_id(prefix):
    return f"{prefix}-{uuid.uuid4().hex}"


def placeholder_image(model, tag="仓位产品"):
    safe_model = text(model) or "未命名"
    safe_tag = text(tag) or "仓位产品"
    svg = f"""
    <svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 600 600">
      <defs>
        <linearGradient id="g" x1="0%" y1="0%" x2="100%" y2="100%">
          <stop offset="0%" stop-color="#e8f1e4"/>
          <stop offset="100%" stop-color="#9bb39b"/>
        </linearGradient>
      </defs>
      <rect width="600" height="600" rx="72" fill="url(#g)"/>
      <rect x="64" y="64" width="472" height="472" rx="48" fill="rgba(255,255,255,0.88)"/>
      <rect x="102" y="108" width="396" height="228" rx="32" fill="rgba(255,255,255,0.94)"/>
      <rect x="102" y="372" width="116" height="80" rx="20" fill="rgba(0,0,0,0.08)"/>
      <rect x="242" y="372" width="116" height="80" rx="20" fill="rgba(0,0,0,0.08)"/>
      <rect x="382" y="372" width="116" height="80" rx="20" fill="rgba(0,0,0,0.08)"/>
      <text x="300" y="208" text-anchor="middle" fill="#334038" font-size="52" font-family="Segoe UI, Arial, sans-serif" font-weight="700">{safe_model}</text>
      <text x="300" y="280" text-anchor="middle" fill="#687268" font-size="28" font-family="Segoe UI, Arial, sans-serif">{safe_tag}</text>
      <text x="160" y="421" text-anchor="middle" fill="#334038" font-size="24" font-family="Segoe UI, Arial, sans-serif">库位</text>
      <text x="300" y="421" text-anchor="middle" fill="#334038" font-size="24" font-family="Segoe UI, Arial, sans-serif">数量</text>
      <text x="440" y="421" text-anchor="middle" fill="#334038" font-size="24" font-family="Segoe UI, Arial, sans-serif">状态</text>
    </svg>
    """
    return f"data:image/svg+xml;charset=UTF-8,{quote(svg)}"


def read_json(path):
    with open(path, "r", encoding="utf-8") as handle:
        return json.load(handle)


def write_json(path, payload):
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)


def write_sheet(workbook, title, headers, rows):
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    sheet.freeze_panes = "A2"
    for column_cells in sheet.columns:
        letter = column_cells[0].column_letter
        width = max(len(text(cell.value)) for cell in column_cells[:24]) if column_cells else 14
        sheet.column_dimensions[letter].width = min(max(width + 2, 12), 32)


def active_rows(rows):
    return [item for item in rows or [] if item.get("status") != "deleted"]


def field_value_map(snapshot):
    result = {}
    for item in active_rows(snapshot.get("productCustomFieldValues", [])):
        product_id = item.get("productId")
        field_id = item.get("fieldId")
        if product_id and field_id:
            result[(product_id, field_id)] = item.get("valueText", "")
    return result


def resolve_balance_position(balance, warehouses_by_id, shelves_by_id, levels_by_id, external_by_id):
    if balance.get("locationType") == "external":
        external = external_by_id.get(balance.get("externalLocationId"))
        return f"非仓库:{external.get('name') or external.get('code')}" if external else ""

    level = levels_by_id.get(balance.get("levelId"))
    shelf = shelves_by_id.get(level.get("shelfId")) if level else None
    warehouse = warehouses_by_id.get(shelf.get("warehouseId")) if shelf else None
    if not warehouse or not shelf or not level:
        return ""
    warehouse_label = warehouse.get("name") or warehouse.get("code")
    return f"{warehouse_label}-{shelf.get('code') or shelf.get('name')}-{level.get('levelNo')}"


def export_import_template(input_path, output_path):
    snapshot = read_json(input_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "导入模板"

    fields = sorted(active_rows(snapshot.get("customFieldDefinitions", [])), key=lambda item: item.get("sortOrder", 0))
    field_headers = [field.get("name", "") for field in fields if field.get("name")]
    headers = ["型号", "图片", "位置", *field_headers]
    sheet.append(headers)

    note_row = ["", "", "按 仓库-货架-层数 填写，如 A仓-S01-1", *["" for _ in field_headers]]
    sheet.append(note_row)

    header_fill = PatternFill("solid", fgColor="DCE8EF")
    note_fill = PatternFill("solid", fgColor="EEF4F0")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="24313A")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    sheet["C2"].font = Font(color="4F6675", italic=True)
    sheet["C2"].fill = note_fill
    sheet["C2"].comment = Comment("位置格式：仓库-货架-层数。只填这个格式即可导入位置，不存在的仓库、货架、层数会自动创建。", "仓位管理")

    sheet.freeze_panes = "A3"
    widths = {
        "A": 22,
        "B": 34,
        "C": 36,
    }
    for letter, width in widths.items():
        sheet.column_dimensions[letter].width = width
    for column_cells in sheet.columns:
        letter = column_cells[0].column_letter
        if letter in widths:
            continue
        width = max(len(text(cell.value)) for cell in column_cells[:24]) if column_cells else 14
        sheet.column_dimensions[letter].width = min(max(width + 2, 12), 24)

    workbook.save(output_path)


def read_sheet_rows(workbook, title):
    if title not in workbook.sheetnames:
      return []

    sheet = workbook[title]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [text(value) for value in rows[0]]
    result = []
    for row in rows[1:]:
        record = {}
        has_value = False
        for index, header in enumerate(headers):
            if not header:
                continue
            value = row[index] if index < len(row) else None
            if value not in (None, ""):
                has_value = True
            record[header] = value
        if has_value:
            result.append(record)
    return result


def split_position(value):
    raw = text(value)
    if not raw:
        return None
    external_match = re.match(r"^(?:非仓库|外部|外仓|external)\s*[:：]\s*(.+)$", raw, re.I)
    if external_match:
        return {"kind": "external", "name": text(external_match.group(1))}
    parts = [text(part) for part in re.split(r"\s*[-－—]\s*", raw) if text(part)]
    if len(parts) < 3:
        return None
    return {
        "kind": "warehouse",
        "warehouse": parts[0],
        "shelf": parts[1],
        "level": "-".join(parts[2:]),
    }


def import_template_workbook(workbook, output_path):
    current_time = now_iso()
    rows = read_sheet_rows(workbook, "导入模板")
    if not rows:
        write_json(output_path, {
            "products": [],
            "customFieldDefinitions": [],
            "productCustomFieldValues": [],
            "warehouses": [],
            "shelves": [],
            "shelfLevels": [],
            "externalLocations": [],
            "inventoryBalances": [],
            "inventoryOperations": [],
            "replaceProductLocations": True,
            "summary": {"products": 0, "balances": 0},
        })
        return

    reserved_headers = {"产品ID", "型号", "产品型号", "图片", "产品图片", "位置", "位置信息", "数量"}
    headers = [header for header in rows[0].keys() if header]
    custom_headers = [header for header in headers if header not in reserved_headers]

    custom_fields = []
    fields_by_name = {}
    for index, header in enumerate(custom_headers, start=1):
        field = {
            "id": make_id("field"),
            "name": header,
            "fieldType": "text",
            "options": [],
            "isRequired": False,
            "isSearchable": True,
            "sortOrder": index,
            "status": "active",
            "createdAt": current_time,
            "updatedAt": current_time,
        }
        custom_fields.append(field)
        fields_by_name[header] = field

    products = []
    products_by_key = {}
    product_values_by_key = {}
    warehouses = []
    warehouses_by_name = {}
    shelves = []
    shelves_by_key = {}
    levels = []
    levels_by_key = {}
    external_locations = []
    external_by_name = {}
    balances_by_key = {}

    def product_key(product_id, model):
        return product_id or normalize(model)

    def ensure_product(row):
        model = text(row.get("型号") or row.get("产品型号"))
        if not model:
            return None
        product_id = text(row.get("产品ID")) or make_id("product")
        key = product_key(text(row.get("产品ID")), model)
        product = products_by_key.get(key) or products_by_key.get(normalize(model))
        image = text(row.get("图片") or row.get("产品图片"))
        if not product:
            product = {
                "id": product_id,
                "model": model,
                "modelNormalized": normalize(model),
                "image": image,
                "status": "active",
                "createdAt": current_time,
                "updatedAt": current_time,
            }
            products.append(product)
            products_by_key[key] = product
            products_by_key[normalize(model)] = product
        else:
            product["model"] = model
            product["modelNormalized"] = normalize(model)
            if image:
                product["image"] = image
            product["updatedAt"] = current_time
        return product

    def ensure_warehouse(name):
        raw = text(name)
        if not raw:
            return None
        key = normalize(raw)
        existing = warehouses_by_name.get(key)
        if existing:
            return existing
        warehouse = {
            "id": make_id("warehouse"),
            "code": raw.upper(),
            "name": raw,
            "colorToken": "blue",
            "sortOrder": len(warehouses) + 1,
            "status": "active",
            "createdAt": current_time,
            "updatedAt": current_time,
        }
        warehouses.append(warehouse)
        warehouses_by_name[key] = warehouse
        return warehouse

    def ensure_shelf(warehouse, code):
        raw = text(code)
        if not warehouse or not raw:
            return None
        key = f"{warehouse['id']}::{normalize(raw)}"
        existing = shelves_by_key.get(key)
        if existing:
            return existing
        shelf = {
            "id": make_id("shelf"),
            "warehouseId": warehouse["id"],
            "code": raw.upper(),
            "name": raw,
            "sortOrder": len([item for item in shelves if item["warehouseId"] == warehouse["id"]]) + 1,
            "status": "active",
            "createdAt": current_time,
            "updatedAt": current_time,
        }
        shelves.append(shelf)
        shelves_by_key[key] = shelf
        return shelf

    def ensure_level(warehouse, shelf, level_text):
        level_no = to_int(re.sub(r"\D", "", text(level_text)), 0)
        if not warehouse or not shelf or level_no <= 0:
            return None
        key = f"{shelf['id']}::{level_no}"
        existing = levels_by_key.get(key)
        if existing:
            return existing
        location_code = f"{warehouse['code']}-{shelf['code']}-L{str(level_no).zfill(2)}"
        level = {
            "id": make_id("level"),
            "shelfId": shelf["id"],
            "levelNo": level_no,
            "locationCode": location_code,
            "qrText": location_code,
            "sortOrder": level_no,
            "status": "active",
            "createdAt": current_time,
            "updatedAt": current_time,
        }
        levels.append(level)
        levels_by_key[key] = level
        return level

    def ensure_external(name):
        raw = text(name)
        if not raw:
            return None
        key = normalize(raw)
        existing = external_by_name.get(key)
        if existing:
            return existing
        external = {
            "id": make_id("external"),
            "code": raw.upper(),
            "name": raw,
            "sortOrder": len(external_locations) + 1,
            "status": "active",
            "createdAt": current_time,
            "updatedAt": current_time,
        }
        external_locations.append(external)
        external_by_name[key] = external
        return external

    for row in rows:
        product = ensure_product(row)
        if not product:
            continue

        for header in custom_headers:
            value_text = text(row.get(header))
            field = fields_by_name.get(header)
            if not field:
                continue
            key = f"{product['id']}::{field['id']}"
            product_values_by_key[key] = {
                "id": product_values_by_key.get(key, {}).get("id") or make_id("pcfv"),
                "productId": product["id"],
                "fieldId": field["id"],
                "valueText": value_text,
                "createdAt": product_values_by_key.get(key, {}).get("createdAt") or current_time,
                "updatedAt": current_time,
            }

        position = split_position(row.get("位置") or row.get("位置信息"))
        if not position:
            continue
        qty = to_float(row.get("数量"), 1) or 1

        if position["kind"] == "external":
            external = ensure_external(position["name"])
            if not external:
                continue
            key = f"{product['id']}::external::{external['id']}"
            balances_by_key[key] = {
                "id": balances_by_key.get(key, {}).get("id") or make_id("balance"),
                "productId": product["id"],
                "locationType": "external",
                "levelId": None,
                "externalLocationId": external["id"],
                "qty": qty,
                "updatedAt": current_time,
            }
            continue

        warehouse = ensure_warehouse(position["warehouse"])
        shelf = ensure_shelf(warehouse, position["shelf"])
        level = ensure_level(warehouse, shelf, position["level"])
        if not level:
            continue
        key = f"{product['id']}::warehouse::{level['id']}"
        balances_by_key[key] = {
            "id": balances_by_key.get(key, {}).get("id") or make_id("balance"),
            "productId": product["id"],
            "locationType": "warehouse",
            "levelId": level["id"],
            "externalLocationId": None,
            "qty": qty,
            "updatedAt": current_time,
        }

    payload = {
        "importMode": "simpleTemplate",
        "replaceProductLocations": True,
        "preserveExistingQty": True,
        "products": products,
        "customFieldDefinitions": custom_fields,
        "productCustomFieldValues": list(product_values_by_key.values()),
        "warehouses": warehouses,
        "shelves": shelves,
        "shelfLevels": levels,
        "externalLocations": external_locations,
        "inventoryBalances": list(balances_by_key.values()),
        "inventoryOperations": [],
        "summary": {
            "products": len(products),
            "fields": len(custom_fields),
            "warehouses": len(warehouses),
            "shelves": len(shelves),
            "levels": len(levels),
            "externalLocations": len(external_locations),
            "balances": len(balances_by_key),
            "operations": 0,
        },
    }
    write_json(output_path, payload)


def export_workbook(input_path, output_path):
    snapshot = read_json(input_path)
    workbook = Workbook()
    workbook.remove(workbook.active)

    write_sheet(
        workbook,
        "说明",
        ["说明", "值"],
        [
            {"说明": "用途", "值": "编辑后可直接导入回系统"},
            {"说明": "建议", "值": "优先修改中文列；ID 列可留空，新行会自动创建"},
            {"说明": "导入方式", "值": "页面右上角“导入 Excel”"},
        ],
    )

    write_sheet(
        workbook,
        "产品",
        ["产品ID", "产品型号", "产品图片"],
        [
            {
                "产品ID": item.get("id", ""),
                "产品型号": item.get("model", ""),
                "产品图片": "" if str(item.get("image", "")).startswith("data:image/") and len(str(item.get("image", ""))) > 3000 else item.get("image", ""),
            }
            for item in snapshot.get("products", [])
        ],
    )

    write_sheet(
        workbook,
        "自定义字段",
        ["字段ID", "字段名", "字段类型", "可选值", "参与搜索", "排序"],
        [
            {
                "字段ID": item.get("id", ""),
                "字段名": item.get("name", ""),
                "字段类型": item.get("fieldType", "text"),
                "可选值": ", ".join(item.get("options", []) or []),
                "参与搜索": bool_label(item.get("isSearchable")),
                "排序": item.get("sortOrder", 0),
            }
            for item in snapshot.get("customFieldDefinitions", [])
        ],
    )

    products_by_id = {item.get("id"): item for item in snapshot.get("products", [])}
    fields_by_id = {item.get("id"): item for item in snapshot.get("customFieldDefinitions", [])}

    write_sheet(
        workbook,
        "产品自定义信息",
        ["记录ID", "产品ID", "产品型号", "字段ID", "字段名", "值"],
        [
            {
                "记录ID": item.get("id", ""),
                "产品ID": item.get("productId", ""),
                "产品型号": products_by_id.get(item.get("productId"), {}).get("model", ""),
                "字段ID": item.get("fieldId", ""),
                "字段名": fields_by_id.get(item.get("fieldId"), {}).get("name", ""),
                "值": item.get("valueText", ""),
            }
            for item in snapshot.get("productCustomFieldValues", [])
        ],
    )

    write_sheet(
        workbook,
        "仓库",
        ["仓库ID", "仓库编码", "仓库名称", "颜色", "排序"],
        [
            {
                "仓库ID": item.get("id", ""),
                "仓库编码": item.get("code", ""),
                "仓库名称": item.get("name", ""),
                "颜色": item.get("colorToken", ""),
                "排序": item.get("sortOrder", 0),
            }
            for item in snapshot.get("warehouses", [])
        ],
    )

    warehouses_by_id = {item.get("id"): item for item in snapshot.get("warehouses", [])}
    write_sheet(
        workbook,
        "货架",
        ["货架ID", "仓库ID", "仓库编码", "货架编码", "货架名称", "排序"],
        [
            {
                "货架ID": item.get("id", ""),
                "仓库ID": item.get("warehouseId", ""),
                "仓库编码": warehouses_by_id.get(item.get("warehouseId"), {}).get("code", ""),
                "货架编码": item.get("code", ""),
                "货架名称": item.get("name", ""),
                "排序": item.get("sortOrder", 0),
            }
            for item in snapshot.get("shelves", [])
        ],
    )

    shelves_by_id = {item.get("id"): item for item in snapshot.get("shelves", [])}
    write_sheet(
        workbook,
        "层数",
        ["层数ID", "货架ID", "仓库编码", "货架编码", "层数", "库位编码", "二维码内容", "排序"],
        [
            {
                "层数ID": item.get("id", ""),
                "货架ID": item.get("shelfId", ""),
                "仓库编码": warehouses_by_id.get(shelves_by_id.get(item.get("shelfId"), {}).get("warehouseId"), {}).get("code", ""),
                "货架编码": shelves_by_id.get(item.get("shelfId"), {}).get("code", ""),
                "层数": item.get("levelNo", 0),
                "库位编码": item.get("locationCode", ""),
                "二维码内容": item.get("qrText", ""),
                "排序": item.get("sortOrder", 0),
            }
            for item in snapshot.get("shelfLevels", [])
        ],
    )

    write_sheet(
        workbook,
        "非仓库位置",
        ["位置ID", "位置编码", "位置名称", "排序"],
        [
            {
                "位置ID": item.get("id", ""),
                "位置编码": item.get("code", ""),
                "位置名称": item.get("name", ""),
                "排序": item.get("sortOrder", 0),
            }
            for item in snapshot.get("externalLocations", [])
        ],
    )

    levels_by_id = {item.get("id"): item for item in snapshot.get("shelfLevels", [])}
    external_by_id = {item.get("id"): item for item in snapshot.get("externalLocations", [])}
    write_sheet(
        workbook,
        "库存",
        ["库存ID", "产品ID", "产品型号", "位置类型", "层数ID", "库位编码", "非仓库位置ID", "非仓库位置编码", "数量"],
        [
            {
                "库存ID": item.get("id", ""),
                "产品ID": item.get("productId", ""),
                "产品型号": products_by_id.get(item.get("productId"), {}).get("model", ""),
                "位置类型": item.get("locationType", ""),
                "层数ID": item.get("levelId", ""),
                "库位编码": levels_by_id.get(item.get("levelId"), {}).get("locationCode", ""),
                "非仓库位置ID": item.get("externalLocationId", ""),
                "非仓库位置编码": external_by_id.get(item.get("externalLocationId"), {}).get("code", ""),
                "数量": item.get("qty", 0),
            }
            for item in snapshot.get("inventoryBalances", [])
        ],
    )

    def level_code(level_id):
        return levels_by_id.get(level_id, {}).get("locationCode", "")

    def external_code(external_id):
        return external_by_id.get(external_id, {}).get("code", "")

    write_sheet(
        workbook,
        "操作",
        [
            "操作ID",
            "批次ID",
            "设备ID",
            "操作类型",
            "产品ID",
            "产品型号",
            "数量",
            "来源类型",
            "来源层数ID",
            "来源库位编码",
            "来源非仓库位置ID",
            "来源非仓库位置编码",
            "目标类型",
            "目标层数ID",
            "目标库位编码",
            "目标非仓库位置ID",
            "目标非仓库位置编码",
            "备注",
            "操作人",
            "操作时间",
            "导入时间",
            "导入状态",
            "设备状态",
        ],
        [
            {
                "操作ID": item.get("id", ""),
                "批次ID": item.get("batchId", ""),
                "设备ID": item.get("deviceId", ""),
                "操作类型": item.get("operationType", ""),
                "产品ID": item.get("productId", ""),
                "产品型号": products_by_id.get(item.get("productId"), {}).get("model", ""),
                "数量": item.get("qty", 0),
                "来源类型": item.get("sourceLocationType", ""),
                "来源层数ID": item.get("sourceLevelId", ""),
                "来源库位编码": level_code(item.get("sourceLevelId")),
                "来源非仓库位置ID": item.get("sourceExternalLocationId", ""),
                "来源非仓库位置编码": external_code(item.get("sourceExternalLocationId")),
                "目标类型": item.get("targetLocationType", ""),
                "目标层数ID": item.get("targetLevelId", ""),
                "目标库位编码": level_code(item.get("targetLevelId")),
                "目标非仓库位置ID": item.get("targetExternalLocationId", ""),
                "目标非仓库位置编码": external_code(item.get("targetExternalLocationId")),
                "备注": item.get("note", ""),
                "操作人": item.get("operatorName", ""),
                "操作时间": item.get("operatedAt", ""),
                "导入时间": item.get("importedAt", ""),
                "导入状态": item.get("importStatus", ""),
                "设备状态": item.get("deviceStatus", ""),
            }
            for item in snapshot.get("inventoryOperations", [])
        ],
    )

    workbook.save(output_path)


def import_workbook(input_path, output_path):
    workbook = load_workbook(input_path, data_only=True)
    current_time = now_iso()

    if "导入模板" in workbook.sheetnames:
        import_template_workbook(workbook, output_path)
        return

    field_rows = read_sheet_rows(workbook, "自定义字段")
    product_rows = read_sheet_rows(workbook, "产品")
    product_value_rows = read_sheet_rows(workbook, "产品自定义信息")
    warehouse_rows = read_sheet_rows(workbook, "仓库")
    shelf_rows = read_sheet_rows(workbook, "货架")
    level_rows = read_sheet_rows(workbook, "层数")
    external_rows = read_sheet_rows(workbook, "非仓库位置")
    balance_rows = read_sheet_rows(workbook, "库存")
    operation_rows = read_sheet_rows(workbook, "操作")

    custom_fields = []
    fields_by_id = {}
    fields_by_name = {}
    for index, row in enumerate(field_rows, start=1):
        name = text(row.get("字段名"))
        if not name:
            continue
        field_id = text(row.get("字段ID")) or make_id("field")
        field = {
            "id": field_id,
            "name": name,
            "fieldType": text(row.get("字段类型")) or "text",
            "options": [item.strip() for item in text(row.get("可选值")).replace("，", ",").split(",") if item.strip()],
            "isRequired": False,
            "isSearchable": parse_bool(row.get("参与搜索")),
            "sortOrder": to_int(row.get("排序"), index),
            "status": "active",
            "createdAt": current_time,
            "updatedAt": current_time,
        }
        custom_fields.append(field)
        fields_by_id[field_id] = field
        fields_by_name[normalize(name)] = field

    products = []
    products_by_id = {}
    products_by_model = {}
    for row in product_rows:
        model = text(row.get("产品型号"))
        if not model:
            continue
        product_id = text(row.get("产品ID")) or make_id("product")
        image = text(row.get("产品图片"))
        product = {
            "id": product_id,
            "model": model,
            "modelNormalized": normalize(model),
            "image": image,
            "status": "active",
            "createdAt": current_time,
            "updatedAt": current_time,
        }
        products.append(product)
        products_by_id[product_id] = product
        products_by_model[normalize(model)] = product

    warehouses = []
    warehouses_by_id = {}
    warehouses_by_code = {}

    def ensure_warehouse(row):
        code = text(row.get("仓库编码")).upper()
        name = text(row.get("仓库名称")) or code
        if not code and not name:
            return None
        warehouse_id = text(row.get("仓库ID"))
        existing = warehouses_by_id.get(warehouse_id) if warehouse_id else None
        existing = existing or warehouses_by_code.get(normalize(code or name))
        if existing:
            return existing
        warehouse = {
            "id": warehouse_id or make_id("warehouse"),
            "code": code or normalize(name),
            "name": name or code,
            "colorToken": text(row.get("颜色")) or "sage",
            "sortOrder": to_int(row.get("排序"), len(warehouses) + 1),
            "status": "active",
            "createdAt": current_time,
            "updatedAt": current_time,
        }
        warehouses.append(warehouse)
        warehouses_by_id[warehouse["id"]] = warehouse
        warehouses_by_code[normalize(warehouse["code"])] = warehouse
        return warehouse

    for row in warehouse_rows:
        ensure_warehouse(row)

    shelves = []
    shelves_by_id = {}
    shelves_by_key = {}

    def ensure_shelf(row):
        warehouse = None
        warehouse_id = text(row.get("仓库ID"))
        if warehouse_id:
            warehouse = warehouses_by_id.get(warehouse_id)
        if not warehouse and text(row.get("仓库编码")):
            warehouse = warehouses_by_code.get(normalize(row.get("仓库编码")))
        if not warehouse and (text(row.get("仓库编码")) or text(row.get("仓库名称"))):
            warehouse = ensure_warehouse(row)

        shelf_code = text(row.get("货架编码")).upper()
        if not warehouse or not shelf_code:
            return None

        shelf_id = text(row.get("货架ID"))
        key = f"{warehouse['id']}::{normalize(shelf_code)}"
        existing = shelves_by_id.get(shelf_id) if shelf_id else None
        existing = existing or shelves_by_key.get(key)
        if existing:
            return existing

        shelf = {
            "id": shelf_id or make_id("shelf"),
            "warehouseId": warehouse["id"],
            "code": shelf_code,
            "name": text(row.get("货架名称")) or f"{shelf_code} 货架",
            "sortOrder": to_int(row.get("排序"), len(shelves) + 1),
            "status": "active",
            "createdAt": current_time,
            "updatedAt": current_time,
        }
        shelves.append(shelf)
        shelves_by_id[shelf["id"]] = shelf
        shelves_by_key[key] = shelf
        return shelf

    for row in shelf_rows:
        ensure_shelf(row)

    levels = []
    levels_by_id = {}
    levels_by_code = {}

    def ensure_level(row):
        shelf = None
        shelf_id = text(row.get("货架ID"))
        if shelf_id:
            shelf = shelves_by_id.get(shelf_id)
        if not shelf and text(row.get("仓库编码")) and text(row.get("货架编码")):
            warehouse = warehouses_by_code.get(normalize(row.get("仓库编码")))
            if warehouse:
                shelf = shelves_by_key.get(f"{warehouse['id']}::{normalize(row.get('货架编码'))}")
        if not shelf:
            shelf = ensure_shelf(row)

        level_no = to_int(row.get("层数"), 0)
        location_code = text(row.get("库位编码"))
        if not shelf or (level_no <= 0 and not location_code):
            return None

        if not location_code:
            warehouse = warehouses_by_id.get(shelf["warehouseId"])
            location_code = f"{warehouse['code']}-{shelf['code']}-L{str(level_no).zfill(2)}"

        level_id = text(row.get("层数ID"))
        existing = levels_by_id.get(level_id) if level_id else None
        existing = existing or levels_by_code.get(normalize(location_code))
        if existing:
            return existing

        level = {
            "id": level_id or make_id("level"),
            "shelfId": shelf["id"],
            "levelNo": level_no or len([item for item in levels if item["shelfId"] == shelf["id"]]) + 1,
            "locationCode": location_code,
            "qrText": text(row.get("二维码内容")) or location_code,
            "sortOrder": to_int(row.get("排序"), level_no or len(levels) + 1),
            "status": "active",
            "createdAt": current_time,
            "updatedAt": current_time,
        }
        levels.append(level)
        levels_by_id[level["id"]] = level
        levels_by_code[normalize(level["locationCode"])] = level
        return level

    for row in level_rows:
        ensure_level(row)

    external_locations = []
    external_by_id = {}
    external_by_code = {}

    def ensure_external(row):
        code = text(row.get("位置编码") or row.get("非仓库位置编码")).upper()
        name = text(row.get("位置名称")) or code
        if not code and not name:
            return None
        external_id = text(row.get("位置ID") or row.get("非仓库位置ID"))
        existing = external_by_id.get(external_id) if external_id else None
        existing = existing or external_by_code.get(normalize(code or name))
        if existing:
            return existing
        location = {
            "id": external_id or make_id("external"),
            "code": code or normalize(name),
            "name": name or code,
            "sortOrder": to_int(row.get("排序"), len(external_locations) + 1),
            "status": "active",
            "createdAt": current_time,
            "updatedAt": current_time,
        }
        external_locations.append(location)
        external_by_id[location["id"]] = location
        external_by_code[normalize(location["code"])] = location
        return location

    for row in external_rows:
        ensure_external(row)

    product_values = []
    for row in product_value_rows:
        value_text = text(row.get("值"))
        if not value_text:
            continue
        product = None
        product_id = text(row.get("产品ID"))
        if product_id:
            product = products_by_id.get(product_id)
        if not product and text(row.get("产品型号")):
            product = products_by_model.get(normalize(row.get("产品型号")))

        field = None
        field_id = text(row.get("字段ID"))
        if field_id:
            field = fields_by_id.get(field_id)
        if not field and text(row.get("字段名")):
            field = fields_by_name.get(normalize(row.get("字段名")))

        if not product or not field:
            continue

        product_values.append(
            {
                "id": text(row.get("记录ID")) or make_id("pcfv"),
                "productId": product["id"],
                "fieldId": field["id"],
                "valueText": value_text,
                "createdAt": current_time,
                "updatedAt": current_time,
            }
        )

    balances_by_key = {}
    for row in balance_rows:
        qty = to_float(row.get("数量"), 0)
        if qty <= 0:
            continue

        product = None
        product_id = text(row.get("产品ID"))
        if product_id:
            product = products_by_id.get(product_id)
        if not product and text(row.get("产品型号")):
            product = products_by_model.get(normalize(row.get("产品型号")))
        if not product:
            continue

        location_type = text(row.get("位置类型")).lower()
        level = None
        external = None

        if location_type in {"external", "非仓库", "非仓库位置"} or text(row.get("非仓库位置ID")) or text(row.get("非仓库位置编码")):
            external = external_by_id.get(text(row.get("非仓库位置ID")))
            if not external and text(row.get("非仓库位置编码")):
                external = external_by_code.get(normalize(row.get("非仓库位置编码")))
            if not external and text(row.get("非仓库位置编码")):
                external = ensure_external({"位置编码": row.get("非仓库位置编码"), "位置名称": row.get("非仓库位置编码")})
            if not external:
                continue
            location_type = "external"
        else:
            level = levels_by_id.get(text(row.get("层数ID")))
            if not level and text(row.get("库位编码")):
                level = levels_by_code.get(normalize(row.get("库位编码")))
            if not level:
                continue
            location_type = "warehouse"

        balance_key = f"{product['id']}::{location_type}::{level['id'] if level else ''}::{external['id'] if external else ''}"
        current = balances_by_key.get(balance_key)
        if current:
            current["qty"] += qty
            continue

        balances_by_key[balance_key] = {
            "id": text(row.get("库存ID")) or make_id("balance"),
            "productId": product["id"],
            "locationType": location_type,
            "levelId": level["id"] if level else None,
            "externalLocationId": external["id"] if external else None,
            "qty": qty,
            "updatedAt": current_time,
        }

    def resolve_operation_location(row, prefix):
        location_type = text(row.get(f"{prefix}类型")).lower()
        external_id = text(row.get(f"{prefix}非仓库位置ID"))
        external_code = text(row.get(f"{prefix}非仓库位置编码"))
        level_id = text(row.get(f"{prefix}层数ID"))
        level_code = text(row.get(f"{prefix}库位编码"))

        if location_type in {"", "none", "无", "-"} and not (external_id or external_code or level_id or level_code):
            return "none", None, None

        if location_type in {"external", "非仓库", "非仓库位置"} or external_id or external_code:
            external = external_by_id.get(external_id) if external_id else None
            if not external and external_code:
                external = external_by_code.get(normalize(external_code))
            if not external and external_code:
                external = ensure_external({"位置编码": external_code, "位置名称": external_code})
            return "external", None, external["id"] if external else None

        level = levels_by_id.get(level_id) if level_id else None
        if not level and level_code:
            level = levels_by_code.get(normalize(level_code))
        return "warehouse", level["id"] if level else None, None

    operations = []
    for row in operation_rows:
        product = None
        product_id = text(row.get("产品ID"))
        if product_id:
            product = products_by_id.get(product_id)
        if not product and text(row.get("产品型号")):
            product = products_by_model.get(normalize(row.get("产品型号")))
        if not product:
            continue

        source_type, source_level_id, source_external_id = resolve_operation_location(row, "来源")
        target_type, target_level_id, target_external_id = resolve_operation_location(row, "目标")
        operation_type = text(row.get("操作类型")) or "put_in"

        operations.append(
            {
                "id": text(row.get("操作ID")) or make_id("operation"),
                "batchId": text(row.get("批次ID")) or make_id("batch"),
                "deviceId": text(row.get("设备ID")),
                "operationType": operation_type,
                "productId": product["id"],
                "qty": to_float(row.get("数量"), 0),
                "sourceLocationType": source_type,
                "sourceLevelId": source_level_id,
                "sourceExternalLocationId": source_external_id,
                "targetLocationType": target_type,
                "targetLevelId": target_level_id,
                "targetExternalLocationId": target_external_id,
                "note": text(row.get("备注")),
                "operatorName": text(row.get("操作人")),
                "operatedAt": text(row.get("操作时间")) or current_time,
                "importedAt": text(row.get("导入时间")) or current_time,
                "importStatus": text(row.get("导入状态")) or "imported",
                "deviceStatus": text(row.get("设备状态")) or "imported",
            }
        )

    payload = {
        "products": products,
        "customFieldDefinitions": custom_fields,
        "productCustomFieldValues": product_values,
        "warehouses": warehouses,
        "shelves": shelves,
        "shelfLevels": levels,
        "externalLocations": external_locations,
        "inventoryBalances": list(balances_by_key.values()),
        "inventoryOperations": operations,
        "summary": {
            "products": len(products),
            "fields": len(custom_fields),
            "warehouses": len(warehouses),
            "shelves": len(shelves),
            "levels": len(levels),
            "externalLocations": len(external_locations),
            "balances": len(balances_by_key),
            "operations": len(operations),
        },
    }
    write_json(output_path, payload)


def main():
    if len(sys.argv) != 4:
        raise SystemExit("Usage: excel_bridge.py <export|template|import> <input> <output>")

    mode = sys.argv[1]
    input_path = Path(sys.argv[2])
    output_path = Path(sys.argv[3])

    if mode == "export":
        export_workbook(input_path, output_path)
        return
    if mode == "template":
        export_import_template(input_path, output_path)
        return
    if mode == "import":
        import_workbook(input_path, output_path)
        return
    raise SystemExit("Unsupported mode.")


if __name__ == "__main__":
    main()
